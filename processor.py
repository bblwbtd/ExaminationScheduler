import re
from typing import Dict, Tuple, List

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet


class StudentCollege:
    def __init__(self, name):
        self.name = name
        self.student_amount = 0

    def add_student(self):
        self.student_amount += 1


class Clazz:
    def __init__(self, name):
        self.name = name
        self.student_colleges: Dict[str, StudentCollege] = {}

    def __gt__(self, other):
        return (self.name == '重修') or (self.name > other.name)

    def __eq__(self, other):
        return self.name == other.name


class Place:
    def __init__(self, name=''):
        self.name = name
        self.clazzes: Dict[str, Clazz] = {}

    def __lt__(self, other):

        def extract_column(name: str):
            left = name.find("(")
            right = name.find(")")

            if left != -1 and right != -1:
                return [int(i) for i in re.findall(r'\d+', name[left: right])]
            return []

        array1 = extract_column(self.name)
        array2 = extract_column(other.name)

        temp1 = self.name[:self.name.find("(")]
        temp2 = other.name[:other.name.find("(")]

        if (len(array1) > 1 and len(array2) > 1) and temp1 == temp2:
            return max(array1) < max(array2)

        return self.name < other.name

    def __eq__(self, other):
        return self.name == other.name


class Course:
    def __init__(self, name, college):
        self.name = name
        self.college = college
        self.places: Dict[str, Place] = {}


class Session:
    def __init__(self, period):
        self.period = period
        self.courses: Dict[str, Course] = {}


class ExamDate:
    def __init__(self, date):
        self.date = date
        self.sessions: Dict[str, Session] = {}


class Campus:
    def __init__(self, name):
        self.name = name
        self.dates: Dict[str, ExamDate] = {}


def parse_date_and_time(date_and_time: str):
    return tuple(date_and_time.split(" "))


name_index_map = {
    "院系": 3,
    "行政班": 5,
    "课程名称": 8,
    "开课院系": 10,
    "考试开始时间": 11,
    "考试结束时间": 12,
    "校区": 13,
    "考试地点": 15,
    "修读类别": 18
}

short_name_map = {
    "文法学院": "文法",
    "马克思主义学院": "马克思",
    "外国语学院": "外语",
    "艺术学院": "艺术",
    "工商管理学院": "管理",
    "理学院": "理学",
    "资源与土木工程学院": "资土",
    "冶金学院": "冶金",
    "材料科学与工程学院": "材料",
    "机械工程与自动化学院": "机械",
    "信息科学与工程学院": "信息",
    "计算机科学与工程学院": "计算机",
    "软件学院": "软件",
    "医学与生物信息工程学院": "医工",
    "生命科学与健康学院": "生命",
    "江河建筑学院": "建筑",
    "机器人科学与工程学院": "机器人",
    "体育部": "体育",
    "医学与生物信息工程学院（原中荷生物医学与信息工程学院）": "医工"
}


class RowInfo:
    def __init__(self, row: Tuple):

        self.student_college = short_name_map.get(row[name_index_map["院系"]], row[name_index_map["院系"]])
        self.clazz = row[name_index_map["行政班"]]
        self.course = row[name_index_map["课程名称"]]
        self.course_college = short_name_map.get(row[name_index_map["开课院系"]], row[name_index_map["开课院系"]])
        self.date, start_time = parse_date_and_time(row[name_index_map["考试开始时间"]])
        _, end_time = parse_date_and_time(row[name_index_map["考试结束时间"]])
        self.period = f'{start_time}-{end_time}'
        self.campus = row[name_index_map["校区"]]
        self.place = row[name_index_map["考试地点"]]
        self.status = row[name_index_map["修读类别"]]

    def cook_info(self, output):
        if self.status == '':
            return

        campus = output.setdefault(self.campus, Campus(self.campus))
        date = campus.dates.setdefault(self.date, ExamDate(self.date))
        session = date.sessions.setdefault(self.period, Session(self.period))
        course = session.courses.setdefault(self.course, Course(self.course, self.course_college))
        place = course.places.setdefault(self.place, Place(self.place))
        if str.strip(self.status) == '重修':
            clazz = place.clazzes.setdefault('重修', Clazz('重修'))
        else:
            clazz = place.clazzes.setdefault(self.clazz, Clazz(self.clazz))

        student_college = clazz.student_colleges.setdefault(self.student_college, StudentCollege(self.student_college))
        student_college.add_student()


def extract_header_index(sheet: Worksheet):
    for cell in sheet[sheet.min_column:sheet.max_column][0]:
        name_index_map[cell.value] = cell.column - 1


def process_file(input_filepath: str):
    wb = load_workbook(input_filepath)
    cooked_data = {}
    for name in wb.sheetnames:
        sheet: Worksheet = wb[name]
        extract_header_index(sheet)
        for index, row in enumerate(sheet.values):
            if index == 0:
                continue
            try:
                row_info = RowInfo(row)
            except Exception as e:
                print(f"Can't parse row {index + 1} {str(e)}")
                continue
            row_info.cook_info(cooked_data)

    return cooked_data


def add_headers_and_title(sheet: Worksheet):
    headers = ('序号', '考试日期', '考试时间', '考试名称', '开课院系', '考试地点', '班级', '考生人数', '考生院系')
    for index, header in enumerate(headers):
        sheet.cell(2, index + 1, header)

    title = '东北大学考试日程表'
    sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=9)
    cell = sheet['A1']
    cell.value = title
    cell.alignment = Alignment(horizontal="center", vertical="center")


def insert_row(sheet: Worksheet, row: int, date, period, course, course_college, place, clazz, student_amount,
               student_college):
    temp = (row - 2, date, period, course, course_college, place, clazz, student_amount, student_college)
    for index, item in enumerate(temp):
        sheet.cell(row, index + 1, item)


column_width = {
    'A': 10,
    'B': 15,
    'C': 15,
    'D': 30,
    'E': 15,
    'F': 30,
    'G': 30,
    'H': 10,
    'I': 30
}


def adjust_column_style(sheet: Worksheet):
    for key in column_width.keys():
        width = column_width[key]
        sheet.column_dimensions[key].width = width


def save_to_sheet(campus: Campus, sheet: Worksheet):
    add_headers_and_title(sheet)
    adjust_column_style(sheet)

    start_pointer = 3
    end_pointer = 3

    dates = list(campus.dates.keys())
    dates.sort()
    for date in dates:
        date_obj = campus.dates[date]

        sessions = list(date_obj.sessions.keys())
        sessions.sort()
        for session in sessions:
            session_obj = date_obj.sessions[session]

            for course, course_obj in session_obj.courses.items():
                places = list(course_obj.places.values())
                places.sort()

                place_pointer = start_pointer
                student_college_pointer = start_pointer

                temp_student_college = None

                for place in places:
                    place_obj = course_obj.places[place.name]
                    array = list(place_obj.clazzes.values())
                    array.sort()
                    for clazz_obj in array:
                        if clazz_obj.name == "重修":
                            total_amount = 0
                            finial_college_name = []
                            for student_college_obj in clazz_obj.student_colleges.values():
                                finial_college_name.append(student_college_obj.name)
                                total_amount += student_college_obj.student_amount
                            temp_student_college = '+'.join(finial_college_name) + '学院'
                            insert_row(sheet, end_pointer, date, session, course, course_obj.college, place.name,
                                       clazz_obj.name,
                                       total_amount, temp_student_college)
                            student_college_pointer = end_pointer
                            end_pointer += 1
                        else:
                            for student_college_obj in clazz_obj.student_colleges.values():
                                insert_row(sheet, end_pointer, date, session, course, course_obj.college, place.name,
                                           clazz_obj.name,
                                           student_college_obj.student_amount, student_college_obj.name)
                                if temp_student_college is None:
                                    temp_student_college = student_college_obj.name
                                elif temp_student_college != clazz_obj.student_colleges:
                                    sheet.merge_cells(start_column=9, end_column=9, start_row=student_college_pointer,
                                                      end_row=end_pointer - 1)
                                    temp_student_college = student_college_obj.name
                                    student_college_pointer = end_pointer

                                end_pointer += 1

                    if place_pointer != end_pointer - 1:
                        sheet.merge_cells(start_row=place_pointer, end_row=end_pointer - 1, start_column=6,
                                          end_column=6)
                    place_pointer = end_pointer

                sheet.merge_cells(start_row=start_pointer, end_row=end_pointer - 1, start_column=5, end_column=5)
                sheet.merge_cells(start_row=start_pointer, end_row=end_pointer - 1, start_column=4, end_column=4)
                sheet.merge_cells(start_row=start_pointer, end_row=end_pointer - 1, start_column=3, end_column=3)
                if student_college_pointer < end_pointer - 1:
                    sheet.merge_cells(start_row=student_college_pointer, end_row=end_pointer - 1, start_column=9,
                                      end_column=9)

                start_pointer = end_pointer


def save_file(data: Dict[str, Campus], output_filepath: str):
    wb = Workbook()
    wb.create_sheet("南湖校区", 0)
    wb.create_sheet("浑南校区", 1)

    if data.get("南湖校区"):
        save_to_sheet(data.get("南湖校区"), wb.worksheets[0])

    if data.get("浑南校区"):
        save_to_sheet(data.get("浑南校区"), wb.worksheets[1])

    wb.save(output_filepath)
